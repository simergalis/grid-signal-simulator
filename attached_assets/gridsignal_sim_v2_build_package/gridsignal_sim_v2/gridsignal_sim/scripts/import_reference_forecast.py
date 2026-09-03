"""Import the annual hourly reference forecast workbook into the ORM tables.

This is intentionally a standalone storage/import utility.  It does not expose
an API route and does not participate in rolling planning or alert generation.

Run from the ``gridsignal_sim`` project directory:

    python -m scripts.import_reference_forecast

The import is idempotent by dataset_id: an existing dataset is reported and
left unchanged, including its resolved rows.
"""

from __future__ import annotations

import argparse
import asyncio
from datetime import datetime, timezone, time
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from sqlalchemy import func, select

from api.db import _SessionLocal, _engine
from runtime.persistence import (
    Base,
    ReferenceForecastResolved,
    ReferenceForecastScenario,
)


DEFAULT_WORKBOOK = (
    Path(__file__).resolve().parents[5]
    / "attached_assets"
    / "scenario-equinix-sj-2-52wk_1787668671730.xlsx"
)
DEFAULT_DATASET_ID = "equinix-sj-2-52wk-v1"
EXPECTED_HEADERS = (
    "Day",
    "Weekday",
    "Hour",
    "Kubernetes",
    "Slurm",
    "Ray",
    "Total",
)


def _as_int(value: Any, *, field: str, row_number: int) -> int:
    if isinstance(value, bool):
        raise ValueError(f"row {row_number}: {field} must be an integer")
    try:
        converted = int(value)
    except (TypeError, ValueError) as exc:
        raise ValueError(f"row {row_number}: {field} must be an integer") from exc
    if isinstance(value, float) and value != converted:
        raise ValueError(f"row {row_number}: {field} must be an integer")
    return converted


def _hour_of_day(value: Any, *, row_number: int) -> int:
    if isinstance(value, time):
        return value.hour
    if isinstance(value, (int, float)) and 0 <= value < 1:
        return int(round(float(value) * 24)) % 24
    if isinstance(value, str):
        text = value.strip()
        try:
            return int(text.split(":", 1)[0])
        except (ValueError, IndexError) as exc:
            raise ValueError(f"row {row_number}: invalid Hour {value!r}") from exc
    raise ValueError(f"row {row_number}: invalid Hour {value!r}")


def read_annual_hourly_rows(workbook_path: Path) -> list[dict[str, int]]:
    """Read and validate the Annual Hourly Detail sheet into plain records."""
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        sheet_name = "Annual Hourly Detail"
        if sheet_name not in workbook.sheetnames:
            raise ValueError(
                f"workbook is missing required sheet {sheet_name!r}; "
                f"found {workbook.sheetnames!r}"
            )
        sheet = workbook[sheet_name]
        rows = sheet.iter_rows(min_row=4, values_only=True)
        headers = tuple(next(rows))
        if headers != EXPECTED_HEADERS:
            raise ValueError(
                f"{sheet_name!r} header mismatch: expected {EXPECTED_HEADERS!r}, "
                f"found {headers!r}"
            )

        resolved: list[dict[str, int]] = []
        for row_number, row in enumerate(rows, start=5):
            if not any(value is not None and value != "" for value in row):
                continue
            if len(row) < len(EXPECTED_HEADERS):
                raise ValueError(f"row {row_number}: expected seven columns")
            day = _as_int(row[0], field="Day", row_number=row_number)
            hour = _hour_of_day(row[2], row_number=row_number)
            kubernetes = _as_int(
                row[3], field="Kubernetes", row_number=row_number
            )
            slurm = _as_int(row[4], field="Slurm", row_number=row_number)
            ray = _as_int(row[5], field="Ray", row_number=row_number)
            total = _as_int(row[6], field="Total", row_number=row_number)
            if not 1 <= day <= 364:
                raise ValueError(f"row {row_number}: Day must be 1..364")
            if not 0 <= hour <= 23:
                raise ValueError(f"row {row_number}: Hour must be 0..23")
            if min(kubernetes, slurm, ray) < 0:
                raise ValueError(f"row {row_number}: counts must be non-negative")
            if total != kubernetes + slurm + ray:
                raise ValueError(
                    f"row {row_number}: Total does not equal domain counts"
                )
            resolved.append(
                {
                    "day_of_year": day,
                    "hour_of_day": hour,
                    "kubernetes_node_count": kubernetes,
                    "slurm_node_count": slurm,
                    "ray_rack_count": ray,
                }
            )
    finally:
        workbook.close()

    expected_keys = {(day, hour) for day in range(1, 365) for hour in range(24)}
    actual_keys = {(row["day_of_year"], row["hour_of_day"]) for row in resolved}
    if len(resolved) != 364 * 24 or actual_keys != expected_keys:
        raise ValueError(
            f"Annual Hourly Detail must contain exactly 8,736 unique "
            f"day/hour rows; found {len(resolved)} rows and "
            f"{len(actual_keys)} unique keys"
        )
    return resolved


async def import_dataset(
    workbook_path: Path, dataset_id: str, display_name: str
) -> tuple[int, int, bool]:
    rows = read_annual_hourly_rows(workbook_path)

    async with _engine.begin() as connection:
        await connection.run_sync(
            lambda sync_connection: Base.metadata.create_all(
                sync_connection,
                tables=[
                    ReferenceForecastScenario.__table__,
                    ReferenceForecastResolved.__table__,
                ],
                checkfirst=True,
            )
        )

    async with _SessionLocal() as session:
        async with session.begin():
            existing = await session.scalar(
                select(ReferenceForecastScenario).where(
                    ReferenceForecastScenario.dataset_id == dataset_id
                )
            )
            if existing is not None:
                # The importer deliberately does not alter an existing dataset.
                count = int(
                    await session.scalar(
                        select(func.count(ReferenceForecastResolved.id)).where(
                            ReferenceForecastResolved.dataset_id == dataset_id
                        )
                    )
                    or 0
                )
                return 1, count, True

            session.add(
                ReferenceForecastScenario(
                    dataset_id=dataset_id,
                    display_name=display_name,
                    span_days=364,
                    source_filename=workbook_path.name,
                    created_at=datetime.now(timezone.utc),
                )
            )
            await session.flush()
            session.add_all(
                ReferenceForecastResolved(
                    dataset_id=dataset_id,
                    **row,
                )
                for row in rows
            )
    return 1, len(rows), False


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--workbook",
        type=Path,
        default=DEFAULT_WORKBOOK,
        help="path to the uploaded .xlsx workbook",
    )
    parser.add_argument("--dataset-id", default=DEFAULT_DATASET_ID)
    parser.add_argument(
        "--display-name",
        default="Equinix SJ-2 52-Week Reference Forecast v1",
    )
    return parser.parse_args()


async def main() -> None:
    args = parse_args()
    if not args.workbook.is_file():
        raise FileNotFoundError(f"workbook not found: {args.workbook}")
    scenario_count, resolved_count, skipped = await import_dataset(
        args.workbook,
        args.dataset_id,
        args.display_name,
    )
    action = "skipped existing" if skipped else "imported"
    print(f"{action}: dataset_id={args.dataset_id}")
    print(f"reference_forecast_scenario rows: {scenario_count}")
    print(f"reference_forecast_resolved rows: {resolved_count}")


if __name__ == "__main__":
    asyncio.run(main())